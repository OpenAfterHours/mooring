"""mooring_deliver — send a notebook's result tables to a stakeholder ``.xlsx``.

mooring INJECTS this module into ``<workspace>/.mooring/pylib/mooring_deliver.py``
and puts that directory on the marimo kernel's import path (see
:func:`mooring.editor.ensure_runtime_config`), so a notebook can::

    import mooring_deliver as md
    md.reset()                          # start fresh each run
    md.table(summary, "Summary")        # one sheet per call
    md.table(by_region, "By region")

and the **Deliver as Excel** action (or ``mooring deliver <path> --excel``) turns
that into one workbook in the sync-excluded ``.mooring/outbox/``. HTML delivery
suits a manager who reads a chart; a finance stakeholder wants the numbers in
Excel, and this is that last mile.

Unlike its sibling runtimes (``mooring_checks`` / ``mooring_inputs``), what this one
writes is deliberately NOT value-free — a workbook of real numbers is the entire
point. That is exactly why it may only ever write under ``.mooring``, which
:func:`mooring.sync.is_synced_path` excludes on both scan sides: the data cannot
ride a push, and nothing here reaches the AI copilot. The one path mooring supplies
from outside (the target, via ``MOORING_DELIVER_XLSX``) is re-checked against that
directory here, so even a rogue environment cannot redirect the values somewhere
syncable. The value-free RECEIPT this leaves behind is what mooring reads back.

**mooring ships no Excel writer.** A frozen ``.pyz`` has no pip at runtime and the
base install is nine dependencies; the workbook is therefore written by the
NOTEBOOK's own environment (the repo's ``pyproject.toml``/``uv.lock``), using
whichever engine is already there — detected at call time. Both ``xlsxwriter`` and
``openpyxl`` are checked, which also covers analysts who reach Excel through polars'
``write_excel`` (xlsxwriter) or pandas' ``to_excel`` (either): the engine is what
actually has to exist. With neither, ``table()`` prints an actionable message and
returns falsy — it never raises, because losing an artifact is a much smaller harm
than breaking the run that produced the numbers.

**The two engines must not disagree.** They differ on the very values a finance
notebook produces — a NaN is a blank cell to one and a hard error to the other, and a
string starting with ``=`` becomes a LIVE FORMULA on both — so every value is
normalised HERE, once, at :func:`table` time (see :func:`_cell`), and the engines
receive only primitives they treat identically. That is also why a table is validated
BEFORE it joins the workbook: an unwritable value can then never poison a later
flush, and it is what lets a per-table failure be RECORDED rather than silently drop
a sheet. Delivery is all-or-nothing on mooring's side — half the numbers looks
exactly like all of them once forwarded.

The provenance sheet written here is a PLACEHOLDER. A delivery run overwrites it from
mooring's side afterwards (:func:`mooring.workbook.stamp_provenance`), because the
notebook must not be able to author the record that vouches for it.

Standalone by design: it imports only the standard library plus the Excel engine it
finds, and duck-types the table you pass (polars OR pandas OR plain Python), so it
works in the team's locked uv env and in the frozen bundle where mooring itself is
not importable. Do not import mooring here.
"""

from __future__ import annotations

import importlib
import inspect
import json
import math
import os
import re
import tempfile
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path

_STATE_DIR = ".mooring"
_WORKBOOKS_DIRNAME = "workbooks"
_OUTBOX_DIRNAME = "outbox"

# What mooring's Deliver-as-Excel run tells the kernel: where to put the workbook, and
# which notebook the receipt belongs to. DELIBERATELY nothing else — the provenance
# facts used to travel this way, and a cell that rewrote os.environ could then forge
# the record vouching for it. They are stamped by mooring after the run instead.
_ENV_TARGET = "MOORING_DELIVER_XLSX"
_ENV_NOTEBOOK = "MOORING_DELIVER_NOTEBOOK"

# Excel's own limits, not ours: 31 characters, and these characters are illegal in a
# sheet name (as is a leading or trailing apostrophe). A workbook that breaks any of
# them is rejected by Excel on open, so the sanitiser below is load-bearing.
_MAX_SHEET_NAME = 31
_BAD_SHEET_CHARS = set("[]:*?/\\")
_PROVENANCE_SHEET = "Provenance"

# A cell holds at most 32,767 characters, and the control characters below are illegal
# in the XML — openpyxl raises on them, xlsxwriter turns them into literal "_x0000_"
# text. Both are normalised here so neither engine decides it on its own.
_MAX_CELL_TEXT = 32767
_TRUNCATED = "…[truncated by mooring]"
_CONTROL_CHARS = re.compile("[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Excel carries 15 significant decimal digits. A wider int or Decimal — an 18-digit
# account number, a full-precision Decimal — is silently rounded by BOTH engines, and
# a rounded account number breaks the join the reader does next, so those are written
# as text: exact, and visibly not a number, rather than quietly wrong.
_EXACT_DIGITS = 15
_EXACT_INT_LIMIT = 10**_EXACT_DIGITS

# One defined rendering for the non-finite floats a ratio produces. openpyxl writes
# them as an EMPTY numeric cell — indistinguishable from blank, from each other, and
# summing as zero — while xlsxwriter refuses the whole workbook. Neither is acceptable
# for a number a stakeholder reads, so both engines get this text instead: never
# blank, never zero, and the same artifact either way.
_NAN_TEXT = "NaN"
_POS_INF_TEXT = "Infinity"
_NEG_INF_TEXT = "-Infinity"

# The one message an analyst with no Excel engine sees. It names the exact command,
# because "pip install openpyxl" is not actionable inside a synced repo — the package
# has to land in the REPO's pyproject.toml for the whole team, which is what
# `mooring deps add` does.
NO_WRITER_HINT = (
    "no Excel writer in this notebook's environment. Add one to the repo with "
    "`mooring deps add openpyxl` (or xlsxwriter) and run Deliver as Excel again. "
    "mooring ships no Excel writer of its own — the workbook is written by the "
    "notebook's environment."
)

# Registered sheets for this kernel, in insertion order: label -> (columns, rows), the
# values already normalised. A dict (not a list) so re-running one cell REPLACES its
# sheet instead of adding a second copy — marimo re-executes cells freely, and a
# duplicated sheet would be a silently wrong deliverable.
_SHEETS: dict[str, tuple[list[str], list[list]]] = {}

# Labels minted for tables passed with no name, keyed by CALL SITE rather than by
# arrival order — so re-running an unnamed md.table(df) replaces its sheet like a named
# one instead of appending "Sheet 2", "Sheet 3", … on every re-execution.
_UNNAMED: dict[tuple[str, int], str] = {}

# What actually reached disk, so a later failure reports the truth ("two sheets are in
# the workbook, the third could not be written") instead of erasing the record of the
# sheets that succeeded.
_DELIVERED: dict[str, object] = {"workbook": "", "sheets": [], "utc_normalised": False}

# Every table that did NOT make it, accumulated for the whole run. A later success must
# never wipe this: mooring refuses a partial delivery, and it can only do that if the
# failure is still on the record when the run ends.
_FAILURES: list[dict] = []


class Result:
    """The outcome of one :func:`table` call. Truthy when the sheet reached the
    workbook on disk; ``repr`` is the one-line summary printed into the cell output."""

    __slots__ = ("name", "rows", "cols", "path", "written", "note")

    def __init__(
        self,
        name: str,
        rows: int = 0,
        cols: int = 0,
        path: str = "",
        written: bool = False,
        note: str = "",
    ) -> None:
        self.name = name
        self.rows = rows
        self.cols = cols
        self.path = path
        self.written = bool(written)
        self.note = note

    def __bool__(self) -> bool:
        return self.written

    def __repr__(self) -> str:
        if not self.written:
            return f"[NOT DELIVERED] sheet {self.name} — {self.note or 'not written'}"
        where = os.path.basename(self.path) if self.path else "the outbox"
        extra = f" ({self.note})" if self.note else ""
        return f"[SHEET] {self.name}{extra} — {self.rows}x{self.cols} -> {where}"


class _WriteError(Exception):
    """Internal: the workbook could not be written. ``reason`` is the mooring-authored
    classification safe to RECORD; ``str(exc)`` may carry engine detail that is only
    ever printed into the analyst's own cell (an engine message can quote a value —
    the same rule that keeps marimo's stderr out of a verify receipt)."""

    def __init__(self, message: str, reason: str = "") -> None:
        super().__init__(message)
        self.reason = reason or message


# -- the public API -------------------------------------------------------------


def table(data=None, name: str | None = None) -> Result:
    """Add ``data`` to the stakeholder workbook as one sheet called ``name``.

    ``data`` is a polars DataFrame/LazyFrame, a pandas DataFrame, a list of dicts, or
    a ``{column: values}`` mapping. Calling ``table`` twice with the same ``name``
    REPLACES that sheet (names match after trimming whitespace, and the printed line
    says "replaced" when it happens), so re-running a cell is idempotent — including
    for an unnamed table, which is keyed by its call site.

    The workbook is rewritten in full on every call, so it is complete and openable
    the moment the last table is registered — there is no ``save()`` to forget.

    Never raises: any failure (no Excel engine, an unreadable table, a locked file)
    prints one line, RECORDS the failure, and returns a falsy :class:`Result`.
    Breaking the run that computed the numbers would be a far worse outcome than
    losing the artifact — and the recorded failure is what makes mooring refuse the
    delivery rather than hand over a workbook quietly missing a sheet."""
    label = _label(name)
    try:
        columns, rows = _extract(data)
        columns = [_text(str(column)) for column in columns]
        rows = [[_cell(value) for value in row] for row in rows]
    except Exception as exc:  # noqa: BLE001  # a bad table must not sink the run
        return _fail(label, str(exc), "could not read the table")
    # Register, write and roll back together: either the sheet is in the workbook on
    # disk or nothing changed. Without that, a sheet the engine rejects stays in the
    # set and takes down every later flush with it — including sheets that were fine.
    snapshot = dict(_SHEETS)
    replaced = label in _SHEETS
    _SHEETS[label] = (columns, rows)
    try:
        path = _flush()
    except _WriteError as exc:
        _SHEETS.clear()
        _SHEETS.update(snapshot)
        return _fail(label, str(exc), exc.reason)
    result = Result(label, len(rows), len(columns), str(path), True, "replaced" if replaced else "")
    print(repr(result))
    return result


def reset() -> None:
    """Clear the registered sheets and remove this notebook's workbook — call at the
    top of the run so a renamed or dropped table cannot linger into today's file.

    Removing the file matters more than clearing the dict: a stale workbook from
    yesterday sitting in the outbox looks exactly like a fresh one to whoever emails
    it, and that is the kind of mistake this product exists to prevent."""
    _SHEETS.clear()
    _UNNAMED.clear()
    _FAILURES.clear()
    _DELIVERED["workbook"] = ""
    _DELIVERED["sheets"] = []
    _DELIVERED["utc_normalised"] = False
    try:
        _target().unlink()
    except OSError:
        pass
    _write_receipt()


# -- table introspection (duck-typed: polars OR pandas OR plain Python) ----------


def _extract(data) -> tuple[list[str], list[list]]:
    """``(columns, rows)`` from whatever the analyst passed.

    Duck-typed in the order that resolves each library unambiguously, and never
    importing polars/pandas — the kernel may have one, both, or neither."""
    if data is None:
        raise ValueError("nothing to write (pass a dataframe, a list of dicts, or a mapping)")
    # A polars LazyFrame: materialise once, then fall through to the DataFrame branch.
    if hasattr(data, "collect") and hasattr(data, "collect_schema"):
        data = data.collect()
    columns = getattr(data, "columns", None)
    if columns is not None and hasattr(data, "rows"):  # polars DataFrame
        return [str(c) for c in columns], [list(r) for r in data.rows()]
    if columns is not None and hasattr(data, "itertuples"):  # pandas DataFrame
        return [str(c) for c in columns], [list(r) for r in data.itertuples(index=False, name=None)]
    if hasattr(data, "keys") and hasattr(data, "values"):  # {column: values}
        names = [str(key) for key in data.keys()]
        cols = [list(values) for values in data.values()]
        height = max((len(col) for col in cols), default=0)
        return names, [[col[i] if i < len(col) else None for col in cols] for i in range(height)]
    try:
        items = list(data)
    except TypeError:
        items = []
    if items and all(hasattr(item, "keys") for item in items):  # a list of dicts
        names = list(dict.fromkeys(str(key) for item in items for key in item.keys()))
        return names, [[item.get(n) for n in names] for item in items]
    raise TypeError(
        "unsupported table: pass a polars/pandas dataframe, a list of dicts, "
        "or a {column: values} mapping"
    )


# -- value normalisation: ONE artifact, whichever engine writes it ---------------


def _cell(value):
    """Normalise one value into a primitive both engines render identically.

    This is where the workbook stops depending on which Excel package the repo happens
    to have. Every branch is here because the engines disagreed, or because Excel
    would have shown the reader something other than the value:

    * **text stays text.** Both engines promote a string starting with ``=`` to a LIVE
      FORMULA, so ``"=1+1"`` reaches the reader as ``2`` — and any upstream free-text
      field an attacker controls (a supplier name, a memo) becomes code in a workbook
      this feature exists to forward outside mooring (``=HYPERLINK`` exfiltration, a
      DDE payload). The write functions below hand strings to the engines' explicit
      TEXT APIs, which is what keeps a cell's value and its appearance the same thing.
    * **non-finite floats become text** (:data:`_NAN_TEXT` and friends): openpyxl
      writes them as an empty numeric cell — a divide-by-zero ratio then reads as "no
      data" and sums as zero — while xlsxwriter refuses the whole workbook.
    * **wide ints and Decimals become text.** Excel carries 15 significant digits;
      wider values are rounded by both engines, and a rounded account number breaks
      the reader's next join. (Both engines DO accept ``Decimal`` and round-trip it
      correctly — the conversion below is about the digit limit, not about support.)
    * **aware datetimes normalise to UTC.** Excel has no timezone concept and both
      engines refuse an aware datetime, so one behaviour has to be chosen: dropping
      the offset and keeping the wall clock lands a single instant on three different
      dates depending on its source offset, which moves a period cut-off."""
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, str):
        return _text(value)
    if isinstance(value, int):
        return value if abs(value) < _EXACT_INT_LIMIT else _text(str(value))
    if isinstance(value, float):
        return value if math.isfinite(value) else _nonfinite_float(value)
    if isinstance(value, Decimal):
        if not value.is_finite():
            return _NAN_TEXT if value.is_nan() else (_POS_INF_TEXT if value > 0 else _NEG_INF_TEXT)
        if len(value.as_tuple().digits) > _EXACT_DIGITS:
            return _text(str(value))
        return float(value)
    if isinstance(value, datetime):
        return _naive_utc(value)
    if isinstance(value, (date, time)):
        return value
    return _text(str(value))


def _nonfinite_float(value: float) -> str:
    if math.isnan(value):
        return _NAN_TEXT
    return _POS_INF_TEXT if value > 0 else _NEG_INF_TEXT


def _naive_utc(value: datetime) -> datetime:
    """An aware datetime converted to UTC and stripped; a naive one left alone. The
    conversion is NOTED so the workbook's provenance can say so — a timestamp column
    whose zone was silently changed is exactly the kind of thing a reader has no way
    to discover afterwards."""
    if value.tzinfo is None:
        return value
    _DELIVERED["utc_normalised"] = True
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _text(value: str) -> str:
    """Drop the control characters Excel's XML forbids and cap the cell's length,
    marking the cut so a truncated memo cannot be mistaken for the whole of one."""
    cleaned = _CONTROL_CHARS.sub("", value)
    if len(cleaned) > _MAX_CELL_TEXT:
        cleaned = cleaned[: _MAX_CELL_TEXT - len(_TRUNCATED)] + _TRUNCATED
    return cleaned


# -- labels and sheet names ------------------------------------------------------


def _label(name: str | None) -> str:
    """The key a table is filed under. A given name wins; an unnamed table is keyed by
    its CALL SITE, so re-running that cell replaces its sheet rather than appending
    another — the same idempotence a named table gets, which a positional default
    would quietly break for exactly the case the optional ``name`` permits."""
    if name is not None and str(name).strip():
        return str(name).strip()
    site = _call_site()
    if site not in _UNNAMED:
        _UNNAMED[site] = f"Sheet {len(_UNNAMED) + 1}"
    return _UNNAMED[site]


def _call_site() -> tuple[str, int]:
    """``(file, line)`` of the first frame outside this module. Best-effort."""
    try:
        for frame_info in inspect.stack(0)[1:]:  # context=0: don't read source lines
            frame = frame_info.frame
            if frame.f_globals.get("__name__") != __name__:
                return str(frame.f_globals.get("__file__", "")), frame.f_lineno
    except Exception:  # noqa: BLE001  # detection is best-effort; never break a run
        pass
    return "", 0


def _sheet_names(labels) -> list[str]:
    """Map the analyst's labels to legal, unique Excel sheet names, in order.

    ``_PROVENANCE_SHEET`` is reserved BEFORE the labels are mapped: mooring's
    provenance record has to be findable under a predictable name (it is rewritten by
    name after the run), so a data sheet the analyst happened to call "Provenance" is
    the one that gets suffixed. Names that collide only after truncation are suffixed
    too — collapsing two sheets into one would drop a table."""
    taken = {_PROVENANCE_SHEET.lower()}
    out = []
    for label in labels:
        base = _legal_sheet_name("".join(" " if c in _BAD_SHEET_CHARS else c for c in label))
        name, n = base, 1
        while name.lower() in taken:
            n += 1
            suffix = f" ({n})"
            name = _legal_sheet_name(base[: _MAX_SHEET_NAME - len(suffix)]) + suffix
        taken.add(name.lower())
        out.append(name)
    return out


def _legal_sheet_name(base: str) -> str:
    """Truncate to Excel's limit and only THEN strip the whitespace and apostrophes it
    forbids at the ends — trimming first lets the cut re-expose a trailing apostrophe,
    which xlsxwriter rejects, failing the whole workbook."""
    return base[:_MAX_SHEET_NAME].strip().strip("'").strip() or "Sheet"


# -- writing the workbook -------------------------------------------------------


def _engine() -> str:
    """The name of the Excel engine available in THIS kernel, or "".

    Probed at call time rather than at import, so an analyst who runs ``mooring deps
    add openpyxl`` mid-session need not restart anything, and so importing
    ``mooring_deliver`` never fails in an environment without a writer."""
    for name in ("xlsxwriter", "openpyxl"):
        try:
            importlib.import_module(name)
        except ImportError:
            continue
        return name
    return ""


def _flush() -> Path:
    """Write every registered sheet plus the provenance placeholder, and return the path.

    Written to a sibling temp file and moved into place, so a reader who opens the
    outbox mid-run never finds a half-written workbook (Excel would call it corrupt)
    and a crashed write leaves the previous file untouched."""
    engine = _engine()
    if not engine:
        raise _WriteError(NO_WRITER_HINT)
    labels = list(_SHEETS)
    names = _sheet_names(labels)
    sheets = [(names[i], *_SHEETS[label]) for i, label in enumerate(labels)]
    sheets.append((_PROVENANCE_SHEET, *_provenance_placeholder(labels)))

    path = _target()
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        handle, tmp = tempfile.mkstemp(
            dir=str(path.parent), prefix="." + path.name + ".", suffix=".tmp"
        )
        os.close(handle)  # the engines open the path themselves
    except OSError as exc:
        raise _WriteError(f"could not create the workbook: {exc}", "could not create it") from exc
    try:
        if engine == "xlsxwriter":
            _write_xlsxwriter(tmp, sheets)
        else:
            _write_openpyxl(tmp, sheets)
        os.replace(tmp, path)
    except Exception as exc:  # noqa: BLE001  # an engine error must not sink the run
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise _WriteError(f"could not write the workbook: {exc}", "could not write it") from exc
    _DELIVERED["workbook"] = _relative(path)
    _DELIVERED["sheets"] = names
    _write_receipt()
    return path


def _write_xlsxwriter(path: str, sheets) -> None:
    import xlsxwriter

    # strings_to_formulas/urls off is the documented switch that stops a value starting
    # with "=" (or looking like a URL) being reinterpreted; write_string below is the
    # belt to that braces, since it never reaches the dispatcher that decides.
    book = xlsxwriter.Workbook(
        path,
        {
            "default_date_format": "yyyy-mm-dd",
            "strings_to_formulas": False,
            "strings_to_urls": False,
        },
    )
    try:
        header = book.add_format({"bold": True})
        for name, columns, rows in sheets:
            sheet = book.add_worksheet(name)
            for col, title in enumerate(columns):
                sheet.write_string(0, col, str(title), header)
            for index, row in enumerate(rows, start=1):
                for col, value in enumerate(row):
                    if isinstance(value, str):
                        sheet.write_string(index, col, value)
                    else:
                        sheet.write(index, col, value)
            sheet.freeze_panes(1, 0)  # the header stays put while a reader scrolls
    finally:
        book.close()


def _write_openpyxl(path: str, sheets) -> None:
    import openpyxl
    from openpyxl.styles import Font

    book = openpyxl.Workbook()
    book.remove(book.active)  # drop the default empty sheet
    bold = Font(bold=True)
    for name, columns, rows in sheets:
        sheet = book.create_sheet(title=name)
        for col, title in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col, value=str(title))
            cell.font = bold
            _force_text(cell)
        for index, row in enumerate(rows, start=2):
            for col, value in enumerate(row, start=1):
                cell = sheet.cell(row=index, column=col, value=value)
                if isinstance(value, str):
                    _force_text(cell)
        sheet.freeze_panes = "A2"  # the header stays put while a reader scrolls
    book.save(path)


def _force_text(cell) -> None:
    """Keep a string cell a STRING. openpyxl inspects the value on assignment and
    promotes anything starting with "=" to a formula cell — which is how a data value
    becomes code in the reader's Excel. Resetting the type afterwards leaves the text
    exactly as written and emits no ``<f>`` element."""
    if getattr(cell, "data_type", "s") != "s":
        cell.data_type = "s"


# -- provenance (a placeholder; mooring overwrites it after the run) -------------


def _provenance_placeholder(labels) -> tuple[list[str], list[list]]:
    """The provenance sheet as the KERNEL can honestly write it: the notebook, the day,
    the sheets. Nothing about the repo or the commit — the notebook is the party being
    vouched for, so it is not handed the facts that vouch for it (a cell can rewrite
    ``os.environ``, and did, in review). A delivery run replaces this sheet wholesale
    from mooring's side, :func:`mooring.workbook.stamp_provenance`; what stands here is
    what an analyst sees building a workbook interactively, claiming nothing more."""
    rows = [
        ["Generated by", "mooring"],
        ["Notebook", _notebook_rel()],
        ["Date", f"{datetime.now():%Y-%m-%d}"],
        ["Sheets", ", ".join(labels)],
    ]
    if _DELIVERED["utc_normalised"]:
        rows.append(["Timestamps", "UTC (timezone-aware values were normalised)"])
    return ["Field", "Value"], rows


# -- where things land ----------------------------------------------------------


def _workspace() -> Path | None:
    # <ws>/.mooring/pylib/mooring_deliver.py -> parents[2] == <ws>
    try:
        return Path(__file__).resolve().parents[2]
    except (OSError, IndexError):
        return None


def _notebook_rel() -> str:
    """The workspace-relative notebook this call belongs to. mooring's own run states it
    outright (so the receipt lands where mooring looks); otherwise it is detected from
    the call stack. This names a FILE, never a provenance claim — mooring re-states the
    notebook itself when it stamps the sheet."""
    given = os.environ.get(_ENV_NOTEBOOK, "").strip().replace("\\", "/")
    if given:
        return given
    ws = _workspace()
    return (_detect_notebook(ws) if ws is not None else None) or "notebook"


def _detect_notebook(ws: Path) -> str | None:
    """The workspace-relative path of the NOTEBOOK that triggered this call.

    marimo sets ``__file__`` in each cell's namespace to the real notebook ``.py``. We
    take the OUTERMOST workspace ``.py`` frame — the notebook cell is the outer frame,
    a helper module it calls is inner — so a table written from a helper is still
    attributed to the notebook that drove it. Best-effort; None outside marimo.
    (Mirrors the same walk in ``mooring_inputs``.)"""
    found: str | None = None
    try:
        for frame_info in inspect.stack(0)[1:]:  # context=0: don't read source lines
            filename = frame_info.frame.f_globals.get("__file__")
            if not isinstance(filename, str) or not filename.endswith(".py"):
                continue
            try:
                rel = Path(filename).resolve().relative_to(ws)
            except (ValueError, OSError):
                continue
            if _STATE_DIR in rel.parts:
                continue  # our own module lives at .mooring/pylib/mooring_deliver.py
            found = str(rel).replace(os.sep, "/")  # keep walking: prefer the OUTERMOST
    except Exception:  # noqa: BLE001  # detection is best-effort; never break a run
        pass
    return found


def _target() -> Path:
    """The workbook's absolute path.

    mooring names it via ``MOORING_DELIVER_XLSX`` so both sides agree on one file
    without re-deriving a date across midnight. That path is CONTAINMENT-CHECKED
    against ``<ws>/.mooring`` before it is honoured: this file carries real data values
    and the sync exclusion is the only thing stopping them riding a push, so an
    environment variable must not be able to move them outside it. Anything else falls
    back to the same dated outbox name the HTML delivery uses."""
    ws = _workspace() or Path.cwd()
    given = os.environ.get(_ENV_TARGET, "").strip()
    if given:
        try:
            candidate = Path(given).resolve()
            if candidate.is_relative_to((ws / _STATE_DIR).resolve()):
                return candidate
        except OSError:
            pass
    rel = _notebook_rel()
    stem = rel.rsplit("/", 1)[-1]
    stem = stem[:-3] if stem.endswith(".py") else stem
    folder = (rel[:-3] if rel.endswith(".py") else rel).replace("/", "__")
    return ws / _STATE_DIR / _OUTBOX_DIRNAME / folder / f"{stem}-{datetime.now():%Y%m%d}.xlsx"


def _relative(path: Path) -> str:
    ws = _workspace()
    if ws is None:
        return str(path)
    try:
        return str(path.relative_to(ws)).replace(os.sep, "/")
    except ValueError:
        return str(path)


# -- the receipt mooring reads back ---------------------------------------------


def _slug(rel: str) -> str:
    """An INJECTIVE per-notebook receipt filename: escape ``_`` first so the ``__``
    that encodes ``/`` is unambiguous (``a/b`` and ``a__b`` map to different files)."""
    return rel.replace("_", "_u").replace("/", "__")


def _fail(label: str, note: str, reason: str) -> Result:
    """Record a table that did not make it, print it, and hand back a falsy Result.

    The recorded ``reason`` is always one of THIS module's own strings, never the
    engine's or the dataframe library's words — those can quote a data value, and the
    receipt is read back by mooring and surfaced in a hub response and on the CLI. The
    detail goes only into ``note``, printed into the analyst's own cell."""
    _FAILURES.append({"sheet": label, "reason": reason})
    _write_receipt()
    result = Result(label, note=note)
    print(repr(result))
    return result


def _write_receipt() -> None:
    """Record what this run delivered and what it could not, so mooring can report it
    without opening the workbook (it has no Excel reader either — see
    :mod:`mooring.workbook`).

    Value-free: the workbook's path, the sheet NAMES the analyst chose, and
    mooring-authored reasons. Failures ACCUMULATE across the run — a later successful
    table must not erase the record of an earlier lost one, because that record is the
    only thing standing between a partial workbook and a stakeholder. Best-effort;
    never raises, since a lost receipt costs a message, not the artifact."""
    ws = _workspace()
    if ws is None:
        return
    rel = _notebook_rel()
    path = ws / _STATE_DIR / _WORKBOOKS_DIRNAME / (_slug(rel) + ".json")
    data = {
        "notebook": rel,
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "workbook": _DELIVERED["workbook"],
        "sheets": list(_DELIVERED["sheets"]),  # type: ignore[arg-type]
        "utc_normalised": bool(_DELIVERED["utc_normalised"]),
        "failures": list(_FAILURES),
        "reason": _FAILURES[0]["reason"] if _FAILURES else "",
    }
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        handle, tmp = tempfile.mkstemp(
            dir=str(path.parent), prefix="." + path.name + ".", suffix=".tmp"
        )
        with os.fdopen(handle, "w", encoding="utf-8", newline="\n") as stream:
            stream.write(json.dumps(data, ensure_ascii=False))
        os.replace(tmp, path)
    except OSError:
        pass
